from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

STATUS_COLORS = {
    "verified": "C6EFCE",
    "discrepancy": "FCE4B2",
    "missing": "F4CCCC",
    "duplicate": "C9DAF8",
}


def write_review_sheet(results: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PURE Review"

    headers = [
        "record_id",
        "title",
        "doi",
        "status",
        "duplicate_score_max",
        "issues",
        "cache_openalex",
        "cache_crossref",
        "cache_scopus",
        "cache_wos",
    ]
    sheet.append(headers)

    for row in results:
        cache = row.get("cache", {})
        issue_str = "; ".join(row.get("issues", []))
        sheet.append(
            [
                row.get("record_id", ""),
                row.get("title", ""),
                row.get("doi", ""),
                row.get("status", ""),
                row.get("duplicate_score_max", 0),
                issue_str,
                cache.get("openalex", ""),
                cache.get("crossref", ""),
                cache.get("scopus", ""),
                cache.get("wos", ""),
            ]
        )

    for row_idx in range(2, sheet.max_row + 1):
        status = str(sheet.cell(row=row_idx, column=4).value or "").lower()
        color = STATUS_COLORS.get(status)
        if not color:
            continue
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col_idx in range(1, len(headers) + 1):
            sheet.cell(row=row_idx, column=col_idx).fill = fill

    for col in sheet.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value or "")
            max_len = max(max_len, len(val))
        sheet.column_dimensions[col_letter].width = min(max_len + 2, 60)

    workbook.save(output_path)
